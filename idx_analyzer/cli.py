"""
Command-line interface for IDX Analyzer
"""

import logging

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

from .analyzer import IDXAnalyzer
from .cache import configure_yfinance_cache, get_cache_info
from .config import Config, create_default_config, load_config
from .exceptions import (
    AnalysisError,
    ChartError,
    IDXAnalyzerError,
    InsufficientDataError,
    InvalidTickerError,
    NetworkError,
    format_error_for_user,
)

logger = logging.getLogger(__name__)

def run_tui_mode():
    """Launch TUI mode"""
    try:
        from .tui.app import run_tui

        run_tui()
    except ImportError as e:
        print(f"Error: TUI dependencies not installed: {e}", file=sys.stderr)
        print("\nTo use TUI mode, ensure textual is installed:", file=sys.stderr)
        print("   uv add textual textual-plotext", file=sys.stderr)
        sys.exit(1)


def create_parser() -> argparse.ArgumentParser:
    """Create argument parser"""
    parser = argparse.ArgumentParser(
        prog="idx-analyzer",
        description="""
 IDX Stock Analyzer v1.0.0
 Indonesian Stock Market Technical Analysis Tool

Analyze Indonesian stocks (IDX) and get support/resistance levels,
trend analysis, and trading recommendations.

Examples:
  idx-analyzer BBCA                    # Analyze BBCA
  idx-analyzer IDX:BBCA                # With IDX: prefix
  idx-analyzer BBCA --period 1y        # 1 year of data
  idx-analyzer TLKM --export csv       # Export to CSV
  idx-analyzer BBCA --chart            # Generate chart
  idx-analyzer ASII -c -p 1y           # Chart with 1 year data
        """,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument(
        "ticker",
        nargs="?",
        help="Stock ticker (e.g., BBCA, TLKM, ASII, or IDX:BBCA). Not required when using --screener.",
    )

    parser.add_argument(
        "-p",
        "--period",
        choices=["1d", "5d", "1mo", "3mo", "6mo", "1y", "2y", "5y"],
        help="Historical data period (default: from config or 6mo)",
    )

    parser.add_argument(
        "-i",
        "--interval",
        choices=["1m", "5m", "15m", "30m", "1h", "1d", "1wk", "1mo"],
        help="Data interval (1m-1h for intraday, limited to 7d-3mo). Default: 1d",
    )

    parser.add_argument(
        "-e", "--export", choices=["csv", "json", "excel"], help="Export analysis to file format"
    )
    parser.add_argument("-o", "--output", help="Output file path for export")

    parser.add_argument(
        "-q",
        "--quiet",
        action="store_true",
        help="Minimal output (useful for scripting)",
    )

    parser.add_argument(
        "-c",
        "--chart",
        action="store_true",
        help="Generate technical analysis chart (PNG)",
    )

    parser.add_argument(
        "--chart-style",
        choices=["standard", "executive"],
        default="standard",
        help="Chart style: 'standard' (classic) or 'executive' (high-end dashboard)",
    )

    parser.add_argument(
        "--chart-output",
        help="Custom chart output filename (default: charts/TICKER/DATE/TICKER_chart.png)",
    )

    parser.add_argument(
        "--patterns",
        action="store_true",
        help="Show candlestick pattern markers on chart",
    )

    parser.add_argument(
        "--macd",
        action="store_true",
        help="Show MACD indicator subplot on chart",
    )

    parser.add_argument(
        "--sentiment-overlay",
        action="store_true",
        help="Overlay sentiment markers on chart (requires --sentiment)",
    )

    parser.add_argument(
        "--all",
        action="store_true",
        help="Enable all chart features: patterns, MACD, sentiment, and sentiment overlay (requires --chart)",
    )

    parser.add_argument(
        "--chat",
        action="store_true",
        help="Generate compact output for Telegram/WhatsApp",
    )

    parser.add_argument(
        "--config",
        help="Path to configuration file",
    )

    parser.add_argument(
        "--init-config",
        action="store_true",
        help="Create default configuration file and exit",
    )

    parser.add_argument(
        "--sentiment",
        action="store_true",
        help="Analyze news sentiment using Yahoo Finance + FinBERT (requires: pip install transformers torch)",
    )

    parser.add_argument(
        "--sentiment-vader",
        action="store_true",
        help="Use lightweight VADER sentiment (no model download, faster)",
    )

    parser.add_argument(
        "--sentiment-llm",
        action="store_true",
        help="Use LLM via OpenAI-compatible API for sentiment (e.g., Ollama)",
    )

    parser.add_argument(
        "--no-hybrid",
        action="store_true",
        help="Disable Indonesian hybrid mode for FinBERT (use pure FinBERT)",
    )

    parser.add_argument(
        "--llm-model",
        default="deepseek-v3.1:671b-cloud",
        help="LLM model name (default: deepseek-v3.1:671b-cloud)",
    )

    parser.add_argument(
        "--llm-url",
        default="http://localhost:11434/v1",
        help="OpenAI-compatible API base URL (default: http://localhost:11434/v1)",
    )

    parser.add_argument(
        "--llm-api-key",
        default="ollama",
        help="API key for the LLM endpoint (default: ollama)",
    )

    parser.add_argument(
        "--cache-info",
        action="store_true",
        help="Show HTTP cache information and exit",
    )

    parser.add_argument(
        "--clear-cache",
        action="store_true",
        help="Clear HTTP cache and exit",
    )

    parser.add_argument(
        "--peers",
        action="store_true",
        help="Show sector peers comparison",
    )

    parser.add_argument(
        "--peers-count",
        type=int,
        default=4,
        help="Number of peers to show (default: 4)",
    )

    parser.add_argument(
        "--peers-compact",
        action="store_true",
        help="Show peers in compact format: (UP 1000 TICKER)",
    )

    # ============================================================================
    # SCREENER OPTIONS
    # ============================================================================
    parser.add_argument(
        "--screener",
        action="store_true",
        help="Run technical stock screener",
    )

    parser.add_argument(
        "--screener-preset",
        choices=["oversold", "overbought", "bullish", "macd", "strong-buy"],
        help="Use preset screener configuration",
    )

    parser.add_argument(
        "--rsi-below",
        type=float,
        metavar="VALUE",
        help="Filter stocks with RSI below VALUE (e.g., 30 for oversold)",
    )

    parser.add_argument(
        "--rsi-above",
        type=float,
        metavar="VALUE",
        help="Filter stocks with RSI above VALUE (e.g., 70 for overbought)",
    )

    parser.add_argument(
        "--above-sma",
        type=int,
        choices=[20, 50, 200],
        metavar="PERIOD",
        help="Filter stocks above SMA period (20, 50, or 200)",
    )

    parser.add_argument(
        "--below-sma",
        type=int,
        choices=[20, 50, 200],
        metavar="PERIOD",
        help="Filter stocks below SMA period (20, 50, or 200)",
    )

    parser.add_argument(
        "--macd-bullish",
        action="store_true",
        help="Filter stocks with bullish MACD (histogram > 0)",
    )

    parser.add_argument(
        "--macd-bearish",
        action="store_true",
        help="Filter stocks with bearish MACD (histogram < 0)",
    )

    parser.add_argument(
        "--change-above",
        type=float,
        metavar="PERCENT",
        help="Filter stocks with daily change above PERCENT%%",
    )

    parser.add_argument(
        "--change-below",
        type=float,
        metavar="PERCENT",
        help="Filter stocks with daily change below PERCENT%%",
    )

    parser.add_argument(
        "--screener-sector",
        metavar="SECTOR",
        help="Screen only stocks in specific sector (e.g., Banking, Mining)",
    )

    parser.add_argument(
        "--screener-index",
        metavar="INDEX",
        help="Screen only stocks in specific index (e.g., LQ45, MSCI, Danantara)",
    )

    parser.add_argument(
        "--screener-board",
        metavar="BOARD",
        help="Screen only stocks in specific board (e.g., Utama, Pengembangan)",
    )

    parser.add_argument(
        "--screener-tickers",
        metavar="TICKERS",
        help="Comma-separated list of tickers to screen (default: liquid universe)",
    )

    parser.add_argument(
        "--screener-workers",
        type=int,
        default=3,
        metavar="N",
        help="Number of concurrent workers for screening (default: 3)",
    )

    parser.add_argument(
        "--screener-export",
        choices=["csv", "json"],
        metavar="FORMAT",
        help="Export screener results to file (csv or json)",
    )

    # --executive is deprecated, use --chart-style executive instead
    parser.add_argument(
        "--executive",
        action="store_true",
        help=argparse.SUPPRESS,  # Hidden from help
    )

    # ============================================================================
    # CORPORATE ACTIONS & BATCH OPTIONS
    # ============================================================================
    parser.add_argument(
        "--dividends",
        action="store_true",
        help="Show dividend history and yield information",
    )

    parser.add_argument(
        "--splits",
        action="store_true",
        help="Show stock splits history",
    )

    parser.add_argument(
        "--actions",
        action="store_true",
        help="Show corporate actions timeline (dividends, splits, earnings)",
    )

    parser.add_argument(
        "--batch",
        metavar="TICKERS",
        help="Batch download data for multiple tickers (comma-separated, e.g., BBCA,BBRI,TLKM)",
    )

    parser.add_argument(
        "--batch-period",
        default="6mo",
        help="Period for batch download (default: 6mo)",
    )

    parser.add_argument("-v", "--version", action="version", version="%(prog)s 1.0.0")

    parser.add_argument(
        "--tui",
        action="store_true",
        help="Launch TUI mode (Bloomberg-style terminal interface)",
    )

    return parser


def get_output_path(base_folder: str, ticker: str, filename: str) -> str:
    """Generate structured output path: base_folder/ticker/YYYY-MM-DD/filename"""
    today = datetime.now().strftime("%Y-%m-%d")
    # Clean ticker just in case
    clean_ticker = ticker.replace("IDX:", "").replace(".JK", "")

    # Structure: base_folder/ticker/date/filename
    target_dir = Path(base_folder) / clean_ticker / today
    target_dir.mkdir(parents=True, exist_ok=True)

    return str(target_dir / filename)


def format_output(result, quiet: bool = False) -> str:
    """Format analysis output"""
    if quiet:
        lines = [
            f"TICKER:{result.ticker}",
            f"PRICE:{result.current_price:.0f}",
            f"CHANGE:{result.change_percent:.2f}",
            f"TREND:{result.trend}",
        ]
        if result.support_levels:
            lines.append(f"SUPPORT:{result.support_levels[0].level:.0f}")
        if result.resistance_levels:
            lines.append(f"RESISTANCE:{result.resistance_levels[0].level:.0f}")
        return "\n".join(lines)

    output = []
    output.append("")
    output.append("=" * 64)
    output.append(f" IDX Stock Analysis: {result.ticker:^42} ")
    output.append("=" * 64)
    output.append("")

    change_color = "+" if result.change_percent >= 0 else "-"
    output.append(f"Current Price: {result.current_price:>12,.0f} IDR")
    output.append(f"   Daily Change:  {change_color} {result.change_percent:>+10.2f}%")
    output.append(f"   Volume:        {result.volume:>12,}")
    output.append("")

    output.append("-" * 64)
    output.append("52-WEEK RANGE")
    output.append("-" * 64)
    from_high = (result.current_price - result.week_52_high) / result.week_52_high * 100
    from_low = (result.current_price - result.week_52_low) / result.week_52_low * 100
    output.append(
        f"   High: {result.week_52_high:>10,.0f}  ({from_high:+.1f}% from current)"
    )
    output.append(
        f"   Low:  {result.week_52_low:>10,.0f}  (+{from_low:.1f}% from current)"
    )
    output.append("")

    output.append("-" * 64)
    output.append("SUPPORT LEVELS (Buy Zones)")
    output.append("-" * 64)
    if result.support_levels:
        for i, s in enumerate(result.support_levels[:4], 1):
            dist = (result.current_price - s.level) / result.current_price * 100
            strength_icon = "*" if s.strength == "strong" else "-"
            output.append(
                f"   {i}. {s.level:>8,.0f}  ({dist:>5.1f}% below)  {strength_icon} {s.strength}"
            )
    else:
        output.append("   No clear support levels identified")
    output.append("")

    output.append("-" * 64)
    output.append("RESISTANCE LEVELS (Sell/Target Zones)")
    output.append("-" * 64)
    if result.resistance_levels:
        for i, r in enumerate(result.resistance_levels[:4], 1):
            dist = (r.level - result.current_price) / result.current_price * 100
            strength_icon = "*" if r.strength == "strong" else "-"
            output.append(
                f"   {i}. {r.level:>8,.0f}  (+{dist:>5.1f}% above)  {strength_icon} {r.strength}"
            )
    else:
        output.append("   No clear resistance levels identified")
    output.append("")

    output.append("-" * 64)
    output.append("MOVING AVERAGES")
    output.append("-" * 64)

    sma_20_pct = (result.current_price - result.sma_20) / result.sma_20 * 100
    sma_50_pct = (result.current_price - result.sma_50) / result.sma_50 * 100

    icon_20 = ">" if result.current_price > result.sma_20 else "<"
    icon_50 = ">" if result.current_price > result.sma_50 else "<"

    output.append(f"   SMA 20: {result.sma_20:>10,.0f}  {icon_20} {sma_20_pct:+.1f}%")
    output.append(f"   SMA 50: {result.sma_50:>10,.0f}  {icon_50} {sma_50_pct:+.1f}%")

    if result.sma_200:
        sma_200_pct = (result.current_price - result.sma_200) / result.sma_200 * 100
        icon_200 = ">" if result.current_price > result.sma_200 else "<"
        output.append(
            f"   SMA 200: {result.sma_200:>9,.0f}  {icon_200} {sma_200_pct:+.1f}%"
        )

    output.append("")

    if result.bb_middle is not None:
        output.append("-" * 64)
        output.append("BOLLINGER BANDS (20, 2)")
        output.append("-" * 64)
        bb_position_text = {
            "above_upper": "Above Upper (Overbought)",
            "near_upper": "Near Upper (Extended)",
            "middle": "Middle (Neutral)",
            "near_lower": "Near Lower (Oversold)",
            "below_lower": "Below Lower (Oversold)",
        }.get(result.bb_position, "Middle")
        output.append(f"   Upper:  {result.bb_upper:>10,.0f}")
        output.append(f"   Middle: {result.bb_middle:>10,.0f}  {bb_position_text}")
        output.append(f"   Lower:  {result.bb_lower:>10,.0f}")
        output.append("")

    if result.vp_poc is not None:
        output.append("-" * 64)
        output.append("VOLUME PROFILE")
        output.append("-" * 64)
        output.append(f"   POC (Point of Control): {result.vp_poc:>10,.0f}")
        output.append(f"   Value Area High:        {result.vp_value_area_high:>10,.0f}")
        output.append(f"   Value Area Low:         {result.vp_value_area_low:>10,.0f}")

        if result.vp_value_area_high and result.vp_value_area_low:
            price_in_va = (
                result.vp_value_area_low
                <= result.current_price
                <= result.vp_value_area_high
            )
            va_status = "Inside Value Area" if price_in_va else "Outside Value Area"
            va_icon = ">" if price_in_va else "~"
            output.append(f"   Current Position:       {va_icon} {va_status}")
        output.append("")

    output.append("-" * 64)
    trend_icon = (
        "BULL"
        if "Bull" in result.trend
        else "BEAR"
        if "Bear" in result.trend
        else "NEUTRAL"
    )
    output.append(f"Trend: {trend_icon} {result.trend}")
    output.append("")
    output.append("RECOMMENDATION:")
    output.append(f"   {result.recommendation}")
    output.append("")
    output.append("-" * 64)
    output.append("")

    return "\n".join(output)


def export_to_csv(result, filepath: str):
    """Export analysis to CSV"""
    import csv

    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Type", "Level", "Distance %", "Strength", "Description"])
        writer.writerow(["Current", result.current_price, "", "", ""])

        for s in result.support_levels:
            dist = (result.current_price - s.level) / result.current_price * 100
            writer.writerow(
                ["Support", s.level, f"{dist:.2f}", s.strength, s.description]
            )

        for r in result.resistance_levels:
            dist = (r.level - result.current_price) / result.current_price * 100
            writer.writerow(
                ["Resistance", r.level, f"{dist:.2f}", r.strength, r.description]
            )

    print(f"Exported to: {filepath}")


def export_to_json(result, filepath: str):
    """Export analysis to JSON"""
    import json

    data = {
        "ticker": result.ticker,
        "current_price": result.current_price,
        "change_percent": result.change_percent,
        "volume": result.volume,
        "week_52_high": result.week_52_high,
        "week_52_low": result.week_52_low,
        "trend": result.trend,
        "recommendation": result.recommendation,
        "support_levels": [
            {
                "level": s.level,
                "strength": s.strength,
                "description": s.description,
                "distance_pct": round(
                    (result.current_price - s.level) / result.current_price * 100, 2
                ),
            }
            for s in result.support_levels
        ],
        "resistance_levels": [
            {
                "level": r.level,
                "strength": r.strength,
                "description": r.description,
                "distance_pct": round(
                    (r.level - result.current_price) / result.current_price * 100, 2
                ),
            }
            for r in result.resistance_levels
        ],
        "bollinger_bands": {
            "middle": result.bb_middle,
            "upper": result.bb_upper,
            "lower": result.bb_lower,
            "position": result.bb_position,
        }
        if result.bb_middle is not None
        else None,
        "volume_profile": {
            "poc": result.vp_poc,
            "value_area_high": result.vp_value_area_high,
            "value_area_low": result.vp_value_area_low,
            "total_volume": result.vp_total_volume,
        }
        if result.vp_poc is not None
        else None,
    }

    with open(filepath, "w") as f:
        json.dump(data, f, indent=2)

    print(f"Exported to: {filepath}")


def export_to_excel(result, filepath: str):
    """Export analysis to Excel with multiple sheets and formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Helper function to style headers
    def style_headers(ws, row_num):
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    # Helper function to auto-adjust column widths
    def auto_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ===== SHEET 1: Summary =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = f"Stock Analysis Report - {result.ticker}"
    ws_summary['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A2'].font = Font(italic=True, size=10)
    ws_summary.merge_cells('A1:D1')
    ws_summary.merge_cells('A2:D2')
    
    # Current Price Section
    ws_summary['A4'] = "Current Price"
    ws_summary['A4'].font = Font(bold=True, size=12)
    ws_summary['B4'] = result.current_price
    ws_summary['B4'].number_format = '#,##0.00'
    ws_summary['B4'].font = Font(bold=True, size=14)
    
    ws_summary['A5'] = "Change %"
    ws_summary['B5'] = result.change_percent / 100
    ws_summary['B5'].number_format = '0.00%'
    if result.change_percent > 0:
        ws_summary['B5'].fill = positive_fill
    elif result.change_percent < 0:
        ws_summary['B5'].fill = negative_fill
    
    ws_summary['A6'] = "Volume"
    ws_summary['B6'] = result.volume
    ws_summary['B6'].number_format = '#,##0'
    
    ws_summary['A7'] = "52-Week High"
    ws_summary['B7'] = result.week_52_high
    ws_summary['B7'].number_format = '#,##0.00'
    
    ws_summary['A8'] = "52-Week Low"
    ws_summary['B8'] = result.week_52_low
    ws_summary['B8'].number_format = '#,##0.00'
    
    ws_summary['A9'] = "Trend"
    ws_summary['B9'] = result.trend
    ws_summary['A10'] = "Recommendation"
    ws_summary['B10'] = result.recommendation
    
    # Summary text
    ws_summary['A12'] = "Summary"
    ws_summary['A12'].font = Font(bold=True, size=12)
    ws_summary['A13'] = result.summary
    ws_summary['A13'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_summary.merge_cells('A13:D15')
    ws_summary.row_dimensions[13].height = 60
    
    # ===== SHEET 2: Technicals =====
    ws_tech = wb.create_sheet("Technicals")
    
    # Technical Indicators
    ws_tech['A1'] = "Technical Indicators"
    ws_tech['A1'].font = Font(bold=True, size=14, color="4472C4")
    
    headers = ["Indicator", "Value", "Status"]
    ws_tech.append(headers)
    style_headers(ws_tech, 2)
    
    # RSI
    rsi_status = "Overbought" if result.rsi > 70 else "Oversold" if result.rsi < 30 else "Neutral"
    ws_tech.append(["RSI (14)", result.rsi, rsi_status])
    
    # Moving Averages
    ws_tech.append(["SMA 20", result.sma_20, ""])
    ws_tech.append(["SMA 50", result.sma_50, ""])
    if result.sma_200:
        ws_tech.append(["SMA 200", result.sma_200, ""])
    
    # MACD
    if result.macd_line:
        macd_status = "Bullish" if result.macd_histogram and result.macd_histogram > 0 else "Bearish"
        ws_tech.append(["MACD Line", result.macd_line, ""])
        ws_tech.append(["MACD Signal", result.macd_signal, ""])
        ws_tech.append(["MACD Histogram", result.macd_histogram, macd_status])
    
    # Bollinger Bands
    if result.bb_middle:
        ws_tech.append(["BB Middle", result.bb_middle, ""])
        ws_tech.append(["BB Upper", result.bb_upper, ""])
        ws_tech.append(["BB Lower", result.bb_lower, ""])
        ws_tech.append(["BB Position", result.bb_position, ""])
    
    # Volume Profile
    if result.vp_poc:
        ws_tech.append(["Volume POC", result.vp_poc, ""])
        ws_tech.append(["Value Area High", result.vp_value_area_high, ""])
        ws_tech.append(["Value Area Low", result.vp_value_area_low, ""])
        ws_tech.append(["Total Volume", result.vp_total_volume, ""])
    
    # Format numbers
    for row in ws_tech.iter_rows(min_row=3, max_row=ws_tech.max_row):
        if isinstance(row[1].value, (int, float)):
            row[1].number_format = '#,##0.00'
        # Apply conditional formatting based on status
        if row[2].value == "Overbought":
            row[2].fill = negative_fill
        elif row[2].value == "Oversold":
            row[2].fill = positive_fill
        elif row[2].value == "Bullish":
            row[2].fill = positive_fill
        elif row[2].value == "Bearish":
            row[2].fill = negative_fill
    
    auto_width(ws_tech)
    
    # Support & Resistance Section
    start_row = ws_tech.max_row + 3
    ws_tech.cell(row=start_row, column=1, value="Support & Resistance Levels")
    ws_tech.cell(row=start_row, column=1).font = Font(bold=True, size=14, color="4472C4")
    
    headers = ["Type", "Level", "Distance %", "Strength", "Description"]
    ws_tech.append(headers)
    style_headers(ws_tech, start_row + 1)
    
    # Support levels
    for s in result.support_levels:
        dist = (result.current_price - s.level) / result.current_price * 100
        ws_tech.append([
            "Support",
            s.level,
            f"-{dist:.2f}%",
            s.strength,
            s.description
        ])
    
    # Resistance levels
    for r in result.resistance_levels:
        dist = (r.level - result.current_price) / result.current_price * 100
        ws_tech.append([
            "Resistance",
            r.level,
            f"+{dist:.2f}%",
            r.strength,
            r.description
        ])
    
    # Format S/R data
    for row in ws_tech.iter_rows(min_row=start_row + 2, max_row=ws_tech.max_row):
        row[1].number_format = '#,##0.00'
        if row[0].value == "Support":
            row[0].fill = positive_fill
        elif row[0].value == "Resistance":
            row[0].fill = negative_fill
    
    auto_width(ws_tech)
    
    # ===== SHEET 3: Fundamentals =====
    ws_fund = wb.create_sheet("Fundamentals")
    
    ws_fund['A1'] = "Fundamental Data"
    ws_fund['A1'].font = Font(bold=True, size=14, color="4472C4")
    
    headers = ["Metric", "Value"]
    ws_fund.append(headers)
    style_headers(ws_fund, 2)
    
    fundamentals = [
        ("Market Cap", result.market_cap, '#,##0'),
        ("P/E Ratio", result.pe_ratio, '0.00'),
        ("Dividend Yield", result.dividend_yield, '0.00%' if result.dividend_yield else None),
    ]
    
    for metric, value, fmt in fundamentals:
        if value is not None:
            row = [metric, value]
            ws_fund.append(row)
            if fmt:
                ws_fund.cell(row=ws_fund.max_row, column=2).number_format = fmt
    
    auto_width(ws_fund)
    
    # Save workbook
    wb.save(filepath)
    print(f"Exported to: {filepath}")


def main(args: Optional[list] = None) -> int:
    """Main entry point"""
    parser = create_parser()
    parsed_args = parser.parse_args(args)

    # Launch TUI mode if requested
    if parsed_args.tui:
        run_tui_mode()
        return 0

    if parsed_args.init_config:
        try:
            config_path = create_default_config()
            print(f"Created default config at: {config_path}")
            return 0
        except Exception as e:
            print(f"Error creating config: {e}", file=sys.stderr)
            return 1

    if parsed_args.cache_info:
        try:
            from .cache import get_cache_info

            info = get_cache_info()
            print("\n" + "=" * 64)
            print(" HTTP CACHE INFORMATION ".center(64))
            print("=" * 64 + "\n")
            print(f"Cache Location:     {info['cache_location']}")
            print(f"Cache Enabled:      {info['cache_enabled']}")
            print(f"Cache TTL:          {info['cache_ttl_hours']:.1f} hours")
            print(f"Cache Exists:       {info['cache_exists']}")
            if "cached_responses" in info:
                print(f"Cached Responses:   {info['cached_responses']}")
            print("\n" + "=" * 64)
            return 0
        except Exception as e:
            print(f"Error getting cache info: {e}", file=sys.stderr)
            return 1

    if parsed_args.clear_cache:
        try:
            from .cache import clear_cache

            clear_cache()
            print("Cache cleared successfully.")
            return 0
        except Exception as e:
            print(f"Error clearing cache: {e}", file=sys.stderr)
            return 1

    config: Optional[Config] = None
    if parsed_args.config:
        try:
            config = load_config(parsed_args.config)
        except Exception as e:
            print(f"Error loading config: {e}", file=sys.stderr)
            return 1

    ticker = parsed_args.ticker

    # ============================================================================
    # SCREENER MODE
    # ============================================================================
    if parsed_args.screener:
        from .screener import (
            TechnicalScreener,
            build_screener_from_args,
            create_bullish_trend_screener,
            create_macd_bullish_screener,
            create_overbought_screener,
            create_oversold_screener,
            create_strong_buy_screener,
        )

        # Build universe from args
        universe = None
        if parsed_args.screener_tickers:
            universe = [
                t.strip().upper() for t in parsed_args.screener_tickers.split(",")
            ]
        elif parsed_args.screener_sector:
            from .stocks_data import get_all_tickers_in_sector

            universe = get_all_tickers_in_sector(parsed_args.screener_sector)
        elif parsed_args.screener_index:
            from .stocks_data import get_index_tickers

            universe = get_index_tickers(parsed_args.screener_index)
        elif parsed_args.screener_board:
            from .stocks_data import get_board_tickers

            universe = get_board_tickers(parsed_args.screener_board)

        # Use preset or build from args
        if parsed_args.screener_preset:
            preset = parsed_args.screener_preset
            if preset == "oversold":
                screener = create_oversold_screener()
            elif preset == "overbought":
                screener = create_overbought_screener()
            elif preset == "bullish":
                screener = create_bullish_trend_screener()
            elif preset == "macd":
                screener = create_macd_bullish_screener()
            elif preset == "strong-buy":
                screener = create_strong_buy_screener()
            else:
                print(f"Unknown preset: {preset}", file=sys.stderr)
                return 1
            # Override universe if specified
            if universe:
                screener.universe = universe
        else:
            screener = build_screener_from_args(parsed_args)
            if universe:
                screener.universe = universe

        # Run screening
        results = screener.screen()

        # Print results
        screener.print_results(only_passed=True)

        # Export if requested
        if parsed_args.screener_export:
            df = screener.to_dataframe(only_passed=True)
            if not df.empty:
                from .screener_export import export_screener_results

                export_screener_results(df, parsed_args.screener_export)

        return 0

    # ============================================================================
    # BATCH DOWNLOAD MODE
    # ============================================================================
    if parsed_args.batch:
        try:
            from .batch_download import batch_download, format_batch_summary

            tickers_list = [t.strip() for t in parsed_args.batch.split(",") if t.strip()]
            if not tickers_list:
                print("Error: --batch requires comma-separated tickers", file=sys.stderr)
                return 1

            print(f"\n📦 Batch downloading {len(tickers_list)} tickers...")
            result = batch_download(
                tickers_list,
                period=parsed_args.batch_period,
            )
            print(format_batch_summary(result))

            if parsed_args.export:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"batch_{timestamp}.{parsed_args.export}"
                output_dir = Path("exports") / "batch"
                output_dir.mkdir(parents=True, exist_ok=True)
                filepath = parsed_args.output or str(output_dir / filename)

                if parsed_args.export == "csv":
                    result.data.to_csv(filepath)
                else:
                    result.data.to_json(filepath, orient="split", indent=2)
                print(f"Exported to: {filepath}")

            return 0
        except Exception as e:
            print(f"Batch download failed: {e}", file=sys.stderr)
            return 1

    # Check ticker is provided for non-screener operations
    if not ticker:
        print(
            "Error: ticker is required (unless using --screener or --batch)",
            file=sys.stderr,
        )
        return 1

    # Initialize HTTP cache for Yahoo Finance API
    try:
        configure_yfinance_cache()
    except Exception as e:
        # Non-fatal: continue without cache
        print(f"Warning: Could not initialize cache: {e}", file=sys.stderr)

    # ============================================================================
    # CORPORATE ACTIONS MODE
    # ============================================================================
    if parsed_args.dividends:
        try:
            from .corporate_actions import format_dividend_report, get_dividend_info

            if not parsed_args.quiet:
                print(f"\nFetching dividend data for {ticker}...")
            info = get_dividend_info(ticker)
            print(format_dividend_report(info))
            return 0
        except Exception as e:
            print(f"Dividend analysis failed: {e}", file=sys.stderr)
            return 1

    if parsed_args.splits:
        try:
            from .corporate_actions import format_splits_report, get_splits_history

            if not parsed_args.quiet:
                print(f"\nFetching stock splits for {ticker}...")
            splits = get_splits_history(ticker)
            print(format_splits_report(ticker, splits))
            return 0
        except Exception as e:
            print(f"Splits history failed: {e}", file=sys.stderr)
            return 1

    if parsed_args.actions:
        try:
            from .corporate_actions import (
                format_corporate_actions_report,
                get_corporate_actions,
            )

            if not parsed_args.quiet:
                print(f"\nFetching corporate actions for {ticker}...")
            actions = get_corporate_actions(ticker)
            print(format_corporate_actions_report(ticker, actions))
            return 0
        except Exception as e:
            print(f"Corporate actions failed: {e}", file=sys.stderr)
            return 1

    if (
        parsed_args.sentiment
        or parsed_args.sentiment_vader
        or parsed_args.sentiment_llm
    ):
        try:
            from .sentiment import SentimentAnalyzer, format_sentiment_report

            use_vader = parsed_args.sentiment_vader
            use_llm = parsed_args.sentiment_llm
            use_hybrid = not parsed_args.no_hybrid
            analyzer = SentimentAnalyzer(
                use_vader=use_vader,
                use_llm=use_llm,
                use_hybrid=use_hybrid,
                llm_model=parsed_args.llm_model,
                llm_base_url=parsed_args.llm_url,
                llm_api_key=parsed_args.llm_api_key,
            )

            if not parsed_args.quiet:
                print(f"\nAnalyzing news sentiment for {ticker}...")
                if use_llm:
                    print(f"   Using LLM: {parsed_args.llm_model}")
                elif not use_vader:
                    if use_hybrid:
                        print(
                            "   Loading FinBERT model with Indonesian hybrid enhancement..."
                        )
                    else:
                        print(
                            "   Loading FinBERT model (first run may take a while)..."
                        )

            result = analyzer.analyze(ticker, max_articles=20)

            if parsed_args.export:
                import json

                if parsed_args.output:
                    filepath = parsed_args.output
                else:
                    filename = f"{ticker}_sentiment.json"
                    filepath = get_output_path("exports", ticker, filename)

                with open(filepath, "w") as f:
                    json.dump(result.to_dict(), f, indent=2)
                print(f"Exported to: {filepath}")
            else:
                print(format_sentiment_report(result))

            return 0

        except ImportError as e:
            print(f"Error: {e}", file=sys.stderr)
            print(
                "\nTo use sentiment analysis, install required packages:",
                file=sys.stderr,
            )
            uv_available = shutil.which("uv") is not None
            if uv_available:
                print("   uv pip install transformers torch", file=sys.stderr)
                print("\nOr run with the 'sentiment' extra:", file=sys.stderr)
                print(
                    "   uv run --extra sentiment idx-analyzer <TICKER> --sentiment",
                    file=sys.stderr,
                )
            else:
                if parsed_args.sentiment_vader:
                    print("   pip install vaderSentiment", file=sys.stderr)
                else:
                    print("   pip install transformers torch", file=sys.stderr)
            return 1
        except Exception as e:
            print(f"Sentiment analysis failed: {e}", file=sys.stderr)
            return 1

    try:
        analyzer = IDXAnalyzer(ticker, config=config)

        period = parsed_args.period
        if period is None and config:
            period = config.analysis.default_period
        elif period is None:
            period = "6mo"

        if not parsed_args.quiet:
            print(f"\nAnalyzing {analyzer.ticker}...")
            print(f"   Fetching {period} of historical data...")

        # Get interval (for intraday support)
        interval = parsed_args.interval
        if interval:
            if not parsed_args.quiet:
                print(f"   Using {interval} interval...")

        analyzer.fetch_data(period=period, interval=interval)
        result = analyzer.analyze()

        # Generate chart (standard or executive style)
        if parsed_args.chart or parsed_args.executive or parsed_args.all:
            # Handle --all flag: enable all chart features
            show_patterns = parsed_args.patterns or parsed_args.all
            show_macd = parsed_args.macd or parsed_args.all
            fetch_sentiment = (
                parsed_args.sentiment
                or parsed_args.sentiment_overlay
                or parsed_args.all
            )
            show_sentiment_overlay = parsed_args.sentiment_overlay or parsed_args.all

            # Handle deprecated --executive flag
            chart_style = (
                "executive" if parsed_args.executive else parsed_args.chart_style
            )
            style_name = (
                "Executive Dashboard" if chart_style == "executive" else "Chart"
            )

            if not parsed_args.quiet:
                print(f"   Generating {style_name.lower()}...")
                if parsed_args.all:
                    print("      (patterns + MACD + sentiment overlay enabled)")

            from .chart import generate_chart

            output_path = parsed_args.chart_output
            if not output_path:
                suffix = "_executive" if chart_style == "executive" else "_chart"
                filename = f"{analyzer.ticker.replace('.JK', '')}{suffix}.png"
                output_path = get_output_path("charts", analyzer.ticker, filename)

            sentiment_data = None
            if show_sentiment_overlay and fetch_sentiment:
                try:
                    from .sentiment import SentimentAnalyzer

                    if not parsed_args.quiet:
                        print("      Fetching news sentiment for overlay...")
                    sent_analyzer = SentimentAnalyzer(use_vader=True)
                    sentiment_result = sent_analyzer.analyze(ticker, max_articles=10)
                    sentiment_data = {
                        "articles": [
                            {
                                "published": str(article.published),
                                "sentiment": article.sentiment_label,
                            }
                            for article in sentiment_result.articles
                        ]
                    }
                except Exception as e:
                    logger.debug(f"Failed to process sentiment data: {e}")
                    pass

            chart_path = generate_chart(
                analyzer=analyzer,
                style=chart_style,
                output_path=output_path,
                show=False,
                show_patterns=show_patterns,
                show_macd=show_macd,
                sentiment_data=sentiment_data,
            )
            if not parsed_args.quiet:
                print(f"{style_name} saved: {chart_path}")

            # If only chart was requested (no other output), return early
            if (
                not parsed_args.export
                and not parsed_args.peers
                and not parsed_args.chat
            ):
                return 0

        if parsed_args.export:
            if parsed_args.output:
                filepath = parsed_args.output
            else:
                ext = "xlsx" if parsed_args.export == "excel" else parsed_args.export
                filename = f"{result.ticker}_analysis.{ext}"
                filepath = get_output_path("exports", result.ticker, filename)

            if parsed_args.export == "csv":
                export_to_csv(result, filepath)
            elif parsed_args.export == "excel":
                export_to_excel(result, filepath)
            else:
                export_to_json(result, filepath)

        if parsed_args.chat:
            print(analyzer.generate_chat_report(result))
            return 0

        # Show sector peers if requested
        if parsed_args.peers:
            if parsed_args.peers_compact:
                from .sector_comparison import format_peer_ticker_style

                peer_line = format_peer_ticker_style(ticker)
                if peer_line:
                    print(f"\n📊 Sector: {peer_line}")
            else:
                from .sector_comparison import format_sector_comparison

                print(
                    "\n"
                    + format_sector_comparison(
                        ticker, max_peers=parsed_args.peers_count
                    )
                )

        if not parsed_args.quiet:
            try:
                from rich.console import Console

                console = Console()
                report = analyzer.generate_rich_report(result)
                console.print(report)
            except ImportError:
                print(format_output(result, quiet=False))
        else:
            print(format_output(result, quiet=True))

        return 0

    except IDXAnalyzerError as e:
        print(format_error_for_user(e), file=sys.stderr)
        return 1
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
